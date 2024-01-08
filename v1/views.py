from django.http import JsonResponse, HttpResponse
from django_ratelimit.decorators import ratelimit
from django.shortcuts import render
from django.utils import timezone

from v1 import BUSINESS_TEXT as btex
from v1.models import Colab, Number
from v1.forms import ColabForm
from v1 import utils

from openpyxl import Workbook


def handle_sms(request):
    phonenumber = request.GET.get("from") # Example : 9148387871
    text = request.GET.get("body")

    if text == '1':
        if utils.number_is_new(phonenumber):
            utils.new_number_row(phonenumber) # save number in db
            utils.send_sms_new_user('0'+phonenumber) # send welcome sms to user
            
        else:
            utils.send_sms_duplicate_number('0'+phonenumber) # this function does nothing 

    else:
        pass # no reaction for anything else only '1'
    
    msg = "Parsa NGO" # This text is only here to keep the api-webhook working
    status_msg = 'ok'
    response = JsonResponse({'status' : status_msg, 'message' : msg}, status=200)
    return response



@ratelimit(key='ip', rate='5/m')
def index_page(request):

    if request.method == "POST":
        form = ColabForm(request.POST)

        if form.is_valid():
            
            # If DB considers data as valid just check phone number like this ->
            if not utils.validate_phonenumber(form.cleaned_data['number']):
                return render(request, "reserve.html", {'msg' : btex.invalid_pnumber_msg})
            
            if not utils.validate_code_melli(form.cleaned_data['code_melli']):
                return render(request, "reserve.html", {'msg' : btex.invalid_cmelli_msg})
            
            if utils.number_is_new(form.cleaned_data['number'][1:]):
                return render(request, "reserve.html", {'msg' : btex.unreg_number_msg})
            
            if not utils.colab_number_is_new(form.cleaned_data['number']):
                return render(request, "reserve.html", {'msg' : btex.duplicate_colab_msg})

            
            try:

                utils.new_colab_row(fullname=form.cleaned_data['fullname'],
                                    number=form.cleaned_data['number'],
                                    code_melli=form.cleaned_data['code_melli'])
                
            except:
                return render(request, "reserve.html", {'msg' : btex.simple_error_msg})
                
            return render(request, "reserve.html", {'msg' : btex.ok_msg})

    else:
        form = ColabForm()

    return render(request, "index.html", {"form": form})


@ratelimit(key='ip', rate='5/m')
def custom_404(request, exception):
    return render(request, '404.html', status=404)

@ratelimit(key='ip', rate='5/m')
def testt_handler(request):
    return render(request, 'reserve.html', {'msg' : btex.wrong_entry_to_res})




def export_colabs(request):
    # Query the database to get the data
    numbers_queryset = Colab.objects.all()

    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Add headers to the sheet
    sheet['A1'] = 'Number'
    sheet['B1'] = 'fullname'
    sheet['C1'] = 'code_melli'

    # Add data to the sheet
    for row_num, number in enumerate(numbers_queryset, start=2):

        sheet.cell(row=row_num, column=1, value=number.number)
        sheet.cell(row=row_num, column=2, value=number.fullname)
        sheet.cell(row=row_num, column=3, value=number.code_melli)


    # Create a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=numbers.xlsx'

    # Save the workbook content to the response
    workbook.save(response)

    return response



def export_numbers(request):
    numbers_queryset = Number.objects.all()

    workbook = Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Number'

    for row_num, number in enumerate(numbers_queryset, start=2):

        sheet.cell(row=row_num, column=1, value=number.number)


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=numbers.xlsx'

    workbook.save(response)

    return response